import os
import time
from dotenv import load_dotenv
import requests
from datetime import datetime
import uuid
import io
import zipfile
import pandas as pd
from dateutil.relativedelta import relativedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from loguru import logger

def DETAIL_HISTORY_REPORT(date: str, TOKEN_NAME: str, max_attempts: int = 5):
    load_dotenv()
    TOKEN_KEY = os.getenv(TOKEN_NAME)
    if TOKEN_KEY == None:
        logger.error('Ошибка инициализации токена!')
        return None
    else:
        logger.info('Токен инициализирован!')
        HEADERS = {'Authorization': TOKEN_KEY}
        url = 'https://seller-analytics-api.wildberries.ru/api/v2/nm-report/downloads'
        report_date =  str(datetime.strptime(date, "%d.%m.%Y").strftime("%Y-%m-%d"))
        new_uuid = str(uuid.uuid4())
        reportType = 'DETAIL_HISTORY_REPORT'
        params = {
            'startDate': report_date,
            'endDate': report_date,
            'skipDeletedNm': False
        }

        body = {
            'id': new_uuid,
            'reportType': reportType,
            'params': params
        }
        response = requests.post(url=url, headers=HEADERS, json=body)
        if response.status_code != 200:
            logger.error(f'Ошибка запроса на формирование отчета за {date} | {response.status_code}')
            return None
        else:
            logger.info(f'Формирование отчета за {date}')
            url = 'https://seller-analytics-api.wildberries.ru/api/v2/nm-report/downloads'
            body = {
                'filter[downloadIds]': [new_uuid]
            }
            
            for attempt in range(max_attempts):
                response = requests.get(url=url, headers=HEADERS, params=body)
                if response.json().get('data')[0].get('status') != 'SUCCESS':
                    if attempt == max_attempts - 1:
                        return None
                    logger.info('Ожидание формирования отчета...')
                    time.sleep(5 * (attempt+1))
                    
                else:
                    logger.success(f'Отчет за {date} готов!')    
                    break
                
            logger.info(f'Получение отчета за {date}')         
            url = f'https://seller-analytics-api.wildberries.ru/api/v2/nm-report/downloads/file/{new_uuid}'
            response = requests.get(url=url, headers=HEADERS)
            
            if response.status_code == 200:
                logger.success(f'Отчет за {date} получен!')
                
                with zipfile.ZipFile(io.BytesIO(response.content), 'r') as zip_file:
                    logger.info(f'Рассчет показателей за {date}')
                    file_content = pd.read_csv(io.StringIO(zip_file.read(f'{new_uuid}.csv').decode('utf-8')), index_col=None)
                    date = file_content['dt'].iloc[0]
                    ordersCount = file_content['ordersCount'].sum() - file_content['cancelCount'].sum()   
                    ordersSum = file_content['ordersSumRub'].sum() - file_content['cancelSumRub'].sum()  
                    buyoutsCount = file_content['buyoutsCount'].sum()
                    buyoutsSum = file_content['buyoutsSumRub'].sum()
                    openCard = file_content['openCardCount'].sum()
                    addToCart = file_content['addToCartCount'].sum()
                    addToCartConversion = file_content.loc[file_content['addToCartConversion'] > 0, 'addToCartConversion'].mean()
                    cartToOrderConversion = file_content.loc[file_content['cartToOrderConversion'] > 0, 'cartToOrderConversion'].mean()
                    buyoutPercent = file_content.loc[file_content['buyoutPercent'] > 0, 'buyoutPercent'].mean()
                    addToWishlist = file_content['addToWishlist'].sum()
                    
                    day_stats = [{
                        'Обновлено': (datetime.now() - relativedelta(hours=3)).strftime('%d.%m.%Y %H:%M'),
                        'Дата': datetime.strptime(date, '%Y-%m-%d').strftime('%d.%m.%Y'),
                        'Заказано, шт.': ordersCount,
                        'Заказано, руб.': ordersSum,
                        'Выкуплено, шт.': buyoutsCount,
                        'Выкуплено, руб.': buyoutsSum,
                        'Переходов в карточки': openCard,
                        'Добавлений в корзину': addToCart,
                        'Средняя конверсия в корзину': addToCartConversion,
                        'Средняя конверсия в заказ': cartToOrderConversion,
                        'Средний процент выкупа': buyoutPercent,
                        'Добавлений в Отложенные': addToWishlist
                    }]
                    
                logger.info(f'Сохранение отчета за {date}')
                filename = 'Ежедневная статистика.xlsx'
                df_new = pd.DataFrame(day_stats)
                file_path = Path(filename)
                
                try:
                    if file_path.exists():
                        df_existing = pd.read_excel(file_path)
                        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                    else:
                        df_combined = df_new
                    df_combined.to_excel(file_path, index=False, engine='openpyxl')
                    wb = load_workbook(filename)
                    ws = wb.active
                        
                    header_font = Font(bold=True, color='000000')
                    header_fill = PatternFill(start_color='CCECFF', fill_type='solid')
                    thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = thin_border
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.border = thin_border
                    for col in ws.columns:
                        max_length = 0
                        column_letter = col[0].column_letter
                        
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 5, 60)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    wb.save(filename)
                    wb.close()
                    logger.success(f"Данные за {date} сохранены в {filename}")
                    
                except PermissionError:
                    logger.error(f'Нет доступа к файлу {filename}')


dates_list = (pd.date_range(end=pd.Timestamp.now(), periods=30, freq='D')).strftime('%d.%m.%Y').tolist()
for date in dates_list:
    DETAIL_HISTORY_REPORT(date,'КОСТРИК')
    time.sleep(30)
