import os
import time
from dotenv import load_dotenv
import requests
from datetime import datetime
import uuid
import io
import zipfile
import pandas as pd
from dateutil.relativedelta import relativedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from loguru import logger
from main import APIRequests


### ------------------------------------------------- ###  

def GET_DAILY_DETAIL_HISTORY_REPORT(TOKEN_NAME: str, date: str, max_attempts: int = 5, DEF_WAIT: int = 15) -> list:
    """
    Функция получения данных из Воронки продаж
    """
    load_dotenv()
    TOKEN_KEY = os.getenv(TOKEN_NAME)
    HEADERS = {'Authorization': TOKEN_KEY}
    url = 'https://seller-analytics-api.wildberries.ru/api/analytics/v3/sales-funnel/products'
    dateStartEnd = str(datetime.strptime(date, "%d.%m.%Y").strftime("%Y-%m-%d"))
    params = {
        'selectedPeriod': {
            'start': dateStartEnd,
            'end': dateStartEnd
        },
        'skipDeletedNm': False,
        'orderBy': {
            'field': 'openCard',
            'mode': 'desc'
        },
        'limit': 1000
    }
    for attempt in range(max_attempts):
        logger.info(f'Получение статистики из Воронки продаж за {date} | попытка {attempt+1}/{max_attempts}')
        response = requests.post(url=url, headers=HEADERS, json=params)
        STATUS_CODE = response.status_code
        if STATUS_CODE == 200:
            products = response.json().get('data').get('products')
            logger.debug(f'Данные из Воронки продаж за {date} получены')
            break   
        elif STATUS_CODE == 429:
            if attempt == max_attempts - 1:
                logger.error(f'{STATUS_CODE} | Превышен лимит запросов. Повторите попытку позже.')
                return None
            else:
                TIME_SLEEP = DEF_WAIT * (attempt+1)
                logger.info(f'{STATUS_CODE} | {response.text} | Ожидание {TIME_SLEEP}сек.')
                time.sleep(TIME_SLEEP)
        else:
            logger.error(f'{STATUS_CODE} | {response.text}')
            return None

    detail_history = []
    for product in products:
        detail_history_product = {}
        detail_history_product['date'] = datetime.strptime(dateStartEnd, '%Y-%m-%d').strftime('%d.%m.%Y')
        detail_history_product['nmId'] = product.get('product').get('nmId')
        detail_history_product['title'] = product.get('product').get('title')
        detail_history_product['vendorCode'] = product.get('product').get('vendorCode')
        detail_history_product['openCount'] = product.get('statistic').get('selected').get('openCount')
        detail_history_product['cartCount'] = product.get('statistic').get('selected').get('cartCount')
        detail_history_product['orderCount'] = product.get('statistic').get('selected').get('orderCount') - product.get('statistic').get('selected').get('cancelCount')
        detail_history_product['orderSum'] = product.get('statistic').get('selected').get('orderSum') - product.get('statistic').get('selected').get('cancelSum')
        detail_history_product['buyoutCount'] = product.get('statistic').get('selected').get('buyoutCount')
        detail_history_product['buyoutSum'] = product.get('statistic').get('selected').get('buyoutSum')
        detail_history_product['addToCartConversion'] = detail_history_product['cartCount'] * 100 / detail_history_product['openCount'] if detail_history_product['openCount'] > 0 else 0
        detail_history_product['cartToOrderConversion'] = detail_history_product['orderCount'] * 100 / detail_history_product['cartCount'] if detail_history_product['cartCount'] > 0 else 0
        detail_history_product['buyoutPercent'] = detail_history_product['buyoutCount'] * 100 / detail_history_product['orderCount'] if detail_history_product['orderCount'] > 0 else 0
        detail_history_product['addToWishlist'] = product.get('statistic').get('selected').get('addToWishlist')
        
        detail_history.append(detail_history_product)
    
    if detail_history:
        logger.success(f'Данные из Воронки продаж за {date} обработаны!')
        return detail_history
    else:
        logger.error(f'Возникла ошибка при обработке данных за {date}')
        return None
    
### ------------------------------------------------- ###  

def CALCULATE_DAILY_DETAIL_HISTORY_REPORT(detail_history: list) -> dict:
    """
    Функция расчета данных по дням из Воронки продаж
    """
    df = pd.DataFrame(detail_history, index=None)
    
    # === Рассчитываем все необходимые показатели ===
    updateDate = datetime.now().strftime('%d.%m.%Y %H:%M')
    reportDate = df['date'].iloc[0]
    orderCount = df['orderCount'].sum()   
    orderSum = df['orderSum'].sum()
    buyoutCount = df['buyoutCount'].sum()
    buyoutSum = df['buyoutSum'].sum()
    showsCount = 0 # Placeholder (Нельзя получить по API)
    openCard = df['openCount'].sum()
    addToCart = df['cartCount'].sum()
    showToClickConversion = 0 # Placeholder (Нельзя посчитать без показов)
    addToCartConversion = (df.loc[df['addToCartConversion'] > 0, 'addToCartConversion'].mean()) / 100
    cartToOrderConversion = (df.loc[df['cartToOrderConversion'] > 0, 'cartToOrderConversion'].mean()) / 100
    buyoutPercent = (df.loc[df['buyoutPercent'] > 0, 'buyoutPercent'].mean()) / 100
    addToWishlist = df['addToWishlist'].sum()
    
    # === Собираем показатели в необходимые поля ===
    day_stats = {
        'updateDate': updateDate,
        'reportDate': reportDate,
        'orderCount': orderCount,
        'orderSum': orderSum,
        'buyoutCount': buyoutCount,
        'buyoutSum': buyoutSum,
        'showsCount': showsCount,
        'openCard': openCard,
        'addToCart': addToCart,
        'showToClickConversion': showToClickConversion,
        'addToCartConversion': addToCartConversion,
        'cartToOrderConversion': cartToOrderConversion,
        'buyoutPercent': buyoutPercent,
        'addToWishlist': addToWishlist
    }
    return day_stats
 
### ------------------------------------------------- ### 
           
def FORMAT_DAILY_DETAIL_HISTORY_REPORT(file: str = 'Ежедневная статистика.xlsx') -> bool | None:
    """
    Функция для форматирования отчета Воронки продаж
    """
    
    # Проверяем существование файла
    file_path = Path(file)
    if not file_path.exists():
        logger.error(f'Файл {file} не найден!')
        return None    
    try:
        logger.info(f'Форматирование файла {file}')
        # === Открываем файл ===
        df = pd.read_excel(file_path)
        
        # === Считаем размер фрейма ===
        last_row = len(df) + 1
        logger.info(f'Всего записей в файле: {last_row-1}')
        
        # === Переименовываем данные для удобства ===
        header_cols = {
        'Обновлено': 'updateDate',
        'Дата': 'reportDate',
        'Заказано, шт.': 'orderCount',
        'Заказано, руб.': 'orderSum',
        'Выкуплено, шт.': 'buyoutCount',
        'Выкуплено, руб.': 'buyoutSum',
        'Показов': 'showsCount',
        'Переходов в карточки': 'openCard',
        'Добавлений в корзину': 'addToCart',
        'Средняя конверсия в клик': 'showToClickConversion',
        'Средняя конверсия в корзину': 'addToCartConversion',
        'Средняя конверсия в заказ': 'cartToOrderConversion',
        'Средний процент выкупа': 'buyoutPercent',
        'Добавлений в Отложенные': 'addToWishlist'
        }
        df = df.rename(columns=header_cols)
        
        # === Сортируем по дате ===
        df['reportDate'] = pd.to_datetime(df['reportDate'], format='%d.%m.%Y', errors='coerce')
        df = df.sort_values('reportDate')
        df['reportDate'] = df['reportDate'].dt.strftime('%d.%m.%Y')
        logger.info(f'Данные отсортированы по дате')
        
        # === Преобразование данных к числам ===
        numeric_cols = {
            'orderCount': int,
            'orderSum': int,
            'buyoutCount': int,
            'buyoutSum': int,
            'showsCount': int,
            'openCard': int,
            'addToCart': int,
            'showToClickConversion': float,
            'addToCartConversion': float,
            'cartToOrderConversion': float,
            'buyoutPercent': float,
            'addToWishlist': int,
        }
        
        for col, dtype in numeric_cols.items():
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                if dtype == int:
                    df[col] = df[col].astype(int)
        logger.info(f'Типы данных преобразованы')
        
        # === Переименовываем столбцы ===
        header_cols = {
            'updateDate': 'Обновлено',
            'reportDate': 'Дата',
            'orderCount': 'Заказано, шт.',
            'orderSum': 'Заказано, руб.',
            'buyoutCount': 'Выкуплено, шт.',
            'buyoutSum': 'Выкуплено, руб.',
            'showsCount': 'Показов',
            'openCard': 'Переходов в карточки',
            'addToCart': 'Добавлений в корзину',
            'showToClickConversion': 'Средняя конверсия в клик',
            'addToCartConversion': 'Средняя конверсия в корзину',
            'cartToOrderConversion': 'Средняя конверсия в заказ',
            'buyoutPercent': 'Средний процент выкупа',
            'addToWishlist': 'Добавлений в Отложенные'
        }
        df = df.rename(columns=header_cols)
        logger.info(f'Столбцы переименованы')

        # === Обработка файла в xlsxwriter ===
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Статистика')
        
        workbook = writer.book
        worksheet = writer.sheets['Статистика']
        
        # === Создание форматов для условного форматирования (xlsxwriter делает это надёжно) ===
        greenFormat = workbook.add_format({'bg_color': "#D1F5D8"})
        redFormat = workbook.add_format({'bg_color': "#FACCD1"})
        
        # === Применяем условное форматирование ===
        for col_letter in ['C', 'D', 'G', 'I', 'J', 'K', 'L']:
            worksheet.conditional_format(
                f'{col_letter}3:{col_letter}{last_row}',
                {'type': 'formula', 'criteria': f'={col_letter}3>{col_letter}2', 'format': greenFormat}
            )
            worksheet.conditional_format(
                f'{col_letter}3:{col_letter}{last_row}',
                {'type': 'formula', 'criteria': f'={col_letter}3<{col_letter}2', 'format': redFormat}
            )
        logger.info(f'Применено условное фомратирование')
        
        # === Автоширина столбцов ===
        for i, col in enumerate(df.columns):
            max_len = max(df[col].fillna(0).astype(str).map(len).max(), len(str(col))) + 2
            worksheet.set_column(i, i, min(max_len, 32))
        logger.info(f'Изменена ширина столбцов')
        
        # === ВАЖНО: Закрываем xlsxwriter ПОЛНОСТЬЮ перед openpyxl ===
        writer.close()
        
        # === Применяем числовые форматы с помощью openpyxl ===
        wb = load_workbook(file_path)
        ws = wb.active
        
        # === Подставим формулу рассчета CTR ===
        for row in range(2, last_row + 1):
            ws[f'J{row}'] = f'=IFERROR(H{row}/G{row},0)'
        logger.info(f'Подставлены формулы')
        
        # === Форматы чисел ===
        rub_format = '#,##0 "₽"'
        count_format = '#,##0'
        percent_format = '0.00%'
        
        # === Форматы по столбцам ===
        column_formats = {
            'A': None,         # Обновлено
            'B': None,         # Дата
            'C': count_format, # Заказано, шт.
            'D': rub_format,   # Заказано, руб.
            'E': count_format, # Выкуплено, шт.
            'F': rub_format,   # Выкуплено, руб.
            'G': count_format, # Показов
            'H': count_format, # Переходов
            'I': count_format, # Добавлений в корзину
            'J': percent_format, # Конверсия в клик
            'K': percent_format, # Конверсия в корзину
            'L': percent_format, # Конверсия в заказ
            'M': percent_format, # Процент выкупа
            'N': count_format, # В отложенные
        }
        
        # === Применяем форматы к ячейкам данных (строки со 2 до last_row) ===
        for col_letter, fmt in column_formats.items():
            if fmt:
                for row in range(2, last_row + 1):
                    ws[f'{col_letter}{row}'].number_format = fmt
        logger.info(f'Форматы чисел изменены')
        
        # === Границы для всех ячеек ===
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        logger.info(f'Границы преобразованы')
        
        # === Форматы заголовка ===
        header_font = Font(bold=True, color='000000')
        header_fill = PatternFill(start_color='CCECFF', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        logger.info(f'Заголовки преобразованы')
        
        # === Сохраняем и закрываем файл ===
        wb.save(file_path)
        wb.close()
        logger.success(f'Файл {file} успешно форматирован и сохранен!')
        return True
    except Exception as e:
        logger.error(f'Произошла ошибка на одном из этапов форматирования файла {file}! | {e}')              
                
### ------------------------------------------------- ###      
      
def UPDATE_DAILY_DETAIL_HISTORY_REPORT(TOKEN_NAME: str, dates: list[str], file_name: str = 'Ежедневная статистика.xlsx'):
    """
    Функция для ежедневного обновления отчета Воронки продаж
    """

    try:
        for date in dates:
            new_daily_stats = GET_DAILY_DETAIL_HISTORY_REPORT(TOKEN_NAME, date)
            new_daily_stats = CALCULATE_DAILY_DETAIL_HISTORY_REPORT(detail_history=new_daily_stats)
                    
            # === Получение сохраненных данных ===
            df_existing = pd.read_excel(file_name)
            
            # === Переименовываем данные для удобства ===
            header_cols = {
            'Обновлено': 'updateDate',
            'Дата': 'reportDate',
            'Заказано, шт.': 'orderCount',
            'Заказано, руб.': 'orderSum',
            'Выкуплено, шт.': 'buyoutCount',
            'Выкуплено, руб.': 'buyoutSum',
            'Показов': 'showsCount',
            'Переходов в карточки': 'openCard',
            'Добавлений в корзину': 'addToCart',
            'Средняя конверсия в клик': 'showToClickConversion',
            'Средняя конверсия в корзину': 'addToCartConversion',
            'Средняя конверсия в заказ': 'cartToOrderConversion',
            'Средний процент выкупа': 'buyoutPercent',
            'Добавлений в Отложенные': 'addToWishlist'
            }
            df_existing = df_existing.rename(columns=header_cols)
            
            # === Вставка новых данных ===
            if new_daily_stats['reportDate'] not in df_existing['reportDate'].values:
                df_new_day_stats = pd.DataFrame([new_daily_stats], index=None)
                df_combined = pd.concat([df_existing, df_new_day_stats], ignore_index=True)
                df_combined.to_excel(file_name, index=False, sheet_name='Статистика')
                logger.success(f"Данные за {date} добавлены в файл {file_name}")
            
            # === Замена устаревших данных ===
            elif new_daily_stats['reportDate'] in df_existing['reportDate'].values:
                # === Выбираем фрейм с датой обрабатываемой на данный момент ===
                df_updatableDate = df_existing[df_existing['reportDate'] == date]
                # === Заменяем данные, если хотя бы какой-то из показателей увеличился ===
                if (new_daily_stats['orderCount'] != df_updatableDate['orderCount'].iloc[0] or 
                    new_daily_stats['orderSum'] != df_updatableDate['orderSum'].iloc[0] or
                    new_daily_stats['buyoutCount'] != df_updatableDate['buyoutCount'].iloc[0] or
                    new_daily_stats['buyoutSum'] != df_updatableDate['buyoutSum'].iloc[0] or
                    new_daily_stats['openCard'] != df_updatableDate['openCard'].iloc[0] or
                    new_daily_stats['addToCart'] != df_updatableDate['addToCart'].iloc[0] or
                    new_daily_stats['addToWishlist'] != df_updatableDate['addToWishlist'].iloc[0]):
                        
                    # === Т.к Показы увеличиться не могут и по API не тянутся, обновлять их не будем ===
                    showsCount = df_updatableDate['showsCount'].iloc[0]
                    showToClickConversion = df_updatableDate['showToClickConversion'].iloc[0]
                    if showsCount > 0 or showToClickConversion > 0:
                        new_daily_stats['showsCount'] = showsCount 
                        new_daily_stats['showToClickConversion'] = showToClickConversion
                        
                    df_new_day_stats = pd.DataFrame([new_daily_stats], index=None)
                    df_existing = df_existing[df_existing['reportDate'] != date]
                    df_combined = pd.concat([df_existing, df_new_day_stats], ignore_index=True)
                    df_combined.to_excel(file_name, index=False, sheet_name='Статистика')
                        
                    logger.success(f"Данные за {date} в файле {file_name} успешно обновлены")
                else:
                    logger.success(f"Данные за {date} в файле {file_name} уже актуальны")
            else:
                logger.critical(f'Неопознанная ошибка!!!')
                
            if date != dates[-1]:
                logger.info(f'Ожидание 20 секунд для предотвращения ошибки 429...')
                time.sleep(20)
        
        if(FORMAT_DAILY_DETAIL_HISTORY_REPORT()):
            logger.success(f'Форматированние данных при обновлении периода {dates[0]} - {dates[-1]} прошло успешно')
        logger.success(f'Данные за период {dates[0]} - {dates[-1]} успешно обновлены')
    
    except Exception as e:
        logger.error(f'Возникла ошибка на одном из этапов обновления данных | {e}')    

### ------------------------------------------------- ###

def GET_REALIZATION_DETAIL_REPORT(TOKEN_NAME: str, date: str, max_attempts: int = 5) -> list:    
    """
    Функция для получения отчета о реализации \n
    date в формате dd.mm.yyyy
    """
    logger.debug(f'Получение отчета о реализации за {date}')
    
    load_dotenv()
    TOKEN_KEY = os.getenv(TOKEN_NAME)
    HEADERS = {'Authorization': TOKEN_KEY}
    
    url = 'https://statistics-api.wildberries.ru/api/v5/supplier/reportDetailByPeriod'
    
    dateFromTo = str(datetime.strptime(date, "%d.%m.%Y").strftime("%Y-%m-%d"))
    params = {
        'dateFrom': dateFromTo,
        'dateTo': dateFromTo,
        'period': 'daily'
    }
    
    for attempt in range(max_attempts):
        response = requests.get(url=url, headers=HEADERS, params=params)
        STATUS_CODE = response.status_code
        if STATUS_CODE == 429:
            WAIT_TIME = 5 * (attempt + 1)
            if attempt == (max_attempts - 1):
                logger.error(f'{STATUS_CODE} | Превышен лимит запросов. Повторите попытку позже.')
                return None
            logger.warning(f'Ошибка {STATUS_CODE}, ожидание {WAIT_TIME} секунд.')
            time.sleep(WAIT_TIME)
        elif STATUS_CODE == 204:
            logger.warning('Нет данных!')
            return None
        elif STATUS_CODE != 200:
            logger.error(f'Ошибка {STATUS_CODE} | {response.text}')
            return None
        else:
            data = response.json()
            logger.success('Все данные получены!')
            break
    
    if not data:
        logger.warning(f'Отчет за {date} пуст !!!')
        return None
    data = response.json()
    logger.success(f'Отчет за {date} получен!')
    
    realization_detail_report = []
    for el in data:
        realization_detail_el = {}
        realization_detail_el['subject_name'] = el.get('subject_name')
        realization_detail_el['nm_id'] = el.get('nm_id')
        realization_detail_el['barcode'] = el.get('barcode')
        realization_detail_el['doc_type_name'] = el.get('doc_type_name')
        realization_detail_el['quantity'] = el.get('quantity')
        realization_detail_el['retail_price'] = el.get('retail_price')
        realization_detail_el['retail_amount'] = el.get('retail_amount')
        realization_detail_el['sale_percent'] = el.get('sale_percent')
        realization_detail_el['commission_percent'] = el.get('commission_percent')
        realization_detail_el['supplier_oper_name'] = el.get('supplier_oper_name')
        realization_detail_el['retail_price_withdisc_rub'] = el.get('retail_price_withdisc_rub')
        realization_detail_el['delivery_amount'] = el.get('delivery_amount')
        realization_detail_el['return_amount'] = el.get('return_amount')
        realization_detail_el['delivery_rub'] = el.get('delivery_rub')
        realization_detail_el['product_discount_for_report'] = el.get('product_discount_for_report')
        realization_detail_el['supplier_promo'] = el.get('supplier_promo')
        realization_detail_el['ppvz_spp_prc'] = el.get('ppvz_spp_prc')
        realization_detail_el['ppvz_kvw_prc_base'] = el.get('ppvz_kvw_prc_base')
        realization_detail_el['ppvz_kvw_prc'] = el.get('ppvz_kvw_prc')
        realization_detail_el['sup_rating_prc_up'] = el.get('sup_rating_prc_up')
        realization_detail_el['is_kgvp_v2'] = el.get('is_kgvp_v2')
        realization_detail_el['ppvz_sales_commission'] = el.get('ppvz_sales_commission')
        realization_detail_el['ppvz_for_pay'] = el.get('ppvz_for_pay')
        realization_detail_el['ppvz_reward'] = el.get('ppvz_reward')
        realization_detail_el['acquiring_fee'] = el.get('acquiring_fee')
        realization_detail_el['acquiring_percent'] = el.get('acquiring_percent')
        realization_detail_el['payment_processing'] = el.get('payment_processing')
        realization_detail_el['ppvz_vw'] = el.get('ppvz_vw')
        realization_detail_el['bonus_type_name'] = el.get('bonus_type_name', None)
        realization_detail_el['penalty'] = el.get('penalty')
        realization_detail_el['additional_payment'] = el.get('additional_payment')
        realization_detail_el['rebill_logistic_cost'] = el.get('rebill_logistic_cost')
        realization_detail_el['storage_fee'] = el.get('storage_fee')
        realization_detail_el['deduction'] = el.get('deduction')
        realization_detail_el['acceptance'] = el.get('acceptance')
        realization_detail_el['srid'] = el.get('srid')
        realization_detail_el['installment_cofinancing_amount'] = el.get('installment_cofinancing_amount')
        realization_detail_el['cashback_amount'] = el.get('cashback_amount')
        realization_detail_el['cashback_discount'] = el.get('cashback_discount')
        realization_detail_el['cashback_commission_change'] = el.get('cashback_commission_change')
        realization_detail_el['order_uid'] = el.get('order_uid')
        realization_detail_el['payment_schedule'] = el.get('payment_schedule')
        realization_detail_el['seller_promo_discount'] = el.get('seller_promo_discount')
        realization_detail_el['loyalty_discount'] = el.get('loyalty_discount')
        realization_detail_el['sale_price_promocode_discount_prc'] = el.get('sale_price_promocode_discount_prc')
        realization_detail_el['sale_price_affiliated_discount_prc'] = el.get('sale_price_affiliated_discount_prc')
        
        realization_detail_report.append(realization_detail_el)
        
    return realization_detail_report
    
### ------------------------------------------------- ### 

def PROCESS_REALIZATION_DETAIL_REPORT(report: list):
    pass
    
### ------------------------------------------------- ### 

def GET_PAID_STORAGE(TOKEN_NAME: str, date: str, max_attempts: int = 5, DEF_WAIT: int = 15) -> list:
    """
    Функция для получения отчета о платном хранении \n
    date в формате dd.mm.yyyy
    """
    logger.debug(f'Получение отчета о платном хранении за {date}')
    
    load_dotenv()
    TOKEN_KEY = os.getenv(TOKEN_NAME)
    HEADERS = {'Authorization': TOKEN_KEY}
    
    url = 'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage'
    
    dateFromTo = str(datetime.strptime(date, "%d.%m.%Y").strftime("%Y-%m-%d"))
    params = {
        'dateFrom': dateFromTo,
        'dateTo': dateFromTo,
    }
    
    logger.info(f'Создание задания на формирование отчета о платном хранении за {date}')
    for attempt in range(max_attempts):
        response = requests.get(url=url, headers=HEADERS, params=params)
        STATUS_CODE = response.status_code
        if STATUS_CODE == 429:
            WAIT_TIME = 5 * (attempt + 1)
            if attempt == (max_attempts - 1):
                logger.error(f'{STATUS_CODE} | Превышен лимит запросов. Повторите попытку позже.')
                return None
            logger.warning(f'Ошибка {STATUS_CODE}, ожидание {WAIT_TIME} секунд.')
            time.sleep(WAIT_TIME)
        elif STATUS_CODE != 200:
            logger.error(f'Ошибка {STATUS_CODE} | {response.text}')
            return None
        else:
            taskId = response.json().get('data').get('taskId')
            logger.debug(f'Задание на генерацию отчета о платном хранении за {date} создано!')
            break
        
    url = f'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{taskId}/status'

    time.sleep(DEF_WAIT)
    logger.info(f'Проверка статуса отчета о платном хранении за {date}')
    for attempt in range(max_attempts):
        response = requests.get(url=url, headers=HEADERS)
        STATUS_CODE = response.status_code
        if STATUS_CODE == 200:
            status = response.json().get('data').get('status')
            if status == 'done':
                break
            elif status == 'new' or status == 'processing':
                WAIT_TIME = DEF_WAIT * (attempt + 1)
                logger.warning(f'Отчет формируется, ожидание {WAIT_TIME} сек.')
                time.sleep(WAIT_TIME)
                continue
            else:
                logger.error(f'Статус отчета {status}, попробуйте повторить позже')
                return None
        elif STATUS_CODE == 429:
            WAIT_TIME = DEF_WAIT * (attempt + 1)
            if attempt == (max_attempts - 1):
                logger.error(f'{STATUS_CODE} | Превышен лимит запросов. Повторите попытку позже.')
                return None
            logger.warning(f'Ошибка {STATUS_CODE}, ожидание {WAIT_TIME} секунд.')
            time.sleep(WAIT_TIME)
        else:
            logger.error(f'Ошибка {STATUS_CODE} | {response.text}')
            return None
        
    url = f'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{taskId}/download'

    logger.info(f'Получение отчета о платном хранении за {date}')
    for attempt in range(max_attempts):
        response = requests.get(url=url, headers=HEADERS)
        STATUS_CODE = response.status_code
        if STATUS_CODE == 200:
            data = response.json()
            logger.debug(f'Отчет о платном хранении за {date} получен!')
            break
        elif STATUS_CODE == 429:
            WAIT_TIME = DEF_WAIT * (attempt + 1)
            if attempt == (max_attempts - 1):
                logger.error(f'{STATUS_CODE} | Превышен лимит запросов. Повторите попытку позже.')
                return None
            logger.warning(f'Ошибка {STATUS_CODE}, ожидание {WAIT_TIME} секунд.')
            time.sleep(WAIT_TIME)
        else:
            logger.error(f'Ошибка {STATUS_CODE} | {response.text}')
            return None
        
    if data:
        logger.info(f'Обработка отчета о платном хранении за {date}')
        paid_storage_df = pd.DataFrame(data, index=None)
        paid_storage_df = (paid_storage_df.groupby(['vendorCode', 'nmId'])['warehousePrice'].sum()).reset_index()
        paid_storage = paid_storage_df.to_dict(orient='records')
        logger.success(f'Отчет о платном хранении за {date} готов!')
        return paid_storage
    else:
        logger.error(f'Ошибка получения отчета о платном хранении за {date}!')
        return None
    
### ------------------------------------------------- ### 

def GET_COST_PRICE(file_path: str = 'required_files/Себестоимости.xlsx') -> list:
    """
    Функция получения себестоимостей из файла Excel
    """
    logger.info(f'Получение себестоимостей')
    try:
        cost_price_df = pd.read_excel(file_path, index_col=None)
        cost_price_df = cost_price_df.rename(columns={
            'Код': 'vendorCode',
            'Наименование': 'title',
            'Себестоимость': 'costPrice',
        })
        cost_price = cost_price_df.to_dict(orient='records')
        logger.success('Себестоимости успешно получены!')
        return cost_price
    except:
        logger.error('Возникла проблема получения себестоимостей товаров!')
        return None

### ------------------------------------------------- ### 

def GET_ADVERTS_LIST(TOKEN_NAME: str, max_attempts: int = 5, DEF_WAIT = 15) -> list:
    """
    Функция возвращает список созданных в кабинете РК
    """
    logger.debug(f'Получение списка рекламных кампаний')
    
    load_dotenv()
    TOKEN_KEY = os.getenv(TOKEN_NAME)
    HEADERS = {'Authorization': TOKEN_KEY}
    
    url = 'https://advert-api.wildberries.ru/adv/v1/promotion/count'
    
    for attempt in range(max_attempts):
        response = requests.get(url=url, headers=HEADERS)
        STATUS_CODE = response.status_code
        if STATUS_CODE == 429:
            WAIT_TIME = DEF_WAIT * (attempt + 1)
            if attempt == (max_attempts - 1):
                logger.error(f'{STATUS_CODE} | Превышен лимит запросов. Повторите попытку позже.')
                return None
            logger.warning(f'Ошибка {STATUS_CODE}, ожидание {WAIT_TIME} секунд.')
            time.sleep(WAIT_TIME)
        elif STATUS_CODE != 200:
            logger.error(f'Ошибка {STATUS_CODE} | {response.text}')
            return None
        else:
            advert_list = response.json().get('adverts')
            advertsIDs = [advert.get('advertId', None) for group in advert_list for advert in group.get('advert_list', [])]
            logger.debug(f'Список РК получен!')
            return advertsIDs

### ------------------------------------------------- ###

def GET_ADVERTS_STATISTIC(TOKEN_NAME: str, date: str, adverts: list = [], max_attempts: int = 5, DEF_WAIT = 15) -> list:
    """
    Функция возвращает статистику всех РК в кабинете
    """
    logger.debug(f'Получение статистики рекламных кампаний')
    
    load_dotenv()
    TOKEN_KEY = os.getenv(TOKEN_NAME)
    HEADERS = {'Authorization': TOKEN_KEY}
    dateFromTo = str(datetime.strptime(date, "%d.%m.%Y").strftime("%Y-%m-%d"))
    
    url = 'https://advert-api.wildberries.ru/adv/v3/fullstats'
    
    chunk_size = 50
    advertsIDs_chunks = [adverts[i:i + chunk_size] for i in range(0, len(adverts), chunk_size)]
    adverts_full_stat = []
    for adverts_chunk in advertsIDs_chunks:
        adverts_chunk_str = ','.join(map(str, adverts_chunk)) 
        params = {
            'ids': adverts_chunk_str,
            'beginDate': dateFromTo,
            'endDate': dateFromTo
        }
        for attempt in range(max_attempts):
            response = requests.get(url=url, headers=HEADERS, params=params)
            STATUS_CODE = response.status_code
            if STATUS_CODE == 429:
                WAIT_TIME = DEF_WAIT * (attempt + 1)
                if attempt == (max_attempts - 1):
                    logger.error(f'{STATUS_CODE} | Превышен лимит запросов. Повторите попытку позже.')
                    return None
                logger.warning(f'Ошибка {STATUS_CODE}, ожидание {WAIT_TIME} секунд.')
                time.sleep(WAIT_TIME)
            elif STATUS_CODE != 200:
                logger.error(f'Ошибка {STATUS_CODE} | {response.text}')
                return None
            else:
                if response.json() is not None:
                    adverts_full_stat += response.json()
                    logger.debug(f'Статистика части РК за {date} добавлена')
                else:
                    logger.debug(f'Статистика части РК за {date} отсутствует')
                break
            
        if adverts_chunk != advertsIDs_chunks[-1]:
            logger.info(f'Ожидание 20 сек. для предотвращения ошибки 429')
            time.sleep(20)
            
    if adverts_full_stat:
        logger.success(f'Полная статистика РК за {date} получена')
        full_stats = []
        for advertId in adverts_full_stat:
            for day in advertId.get('days'):
                for app in day.get('apps'):
                    for nm in app.get('nms'):
                        nm_stats = {}
                        nm_stats['nmId'] = nm.get('nmId')
                        nm_stats['sum'] = nm.get('sum')
                        nm_stats['atbs'] = nm.get('atbs')
                        nm_stats['canceled'] = nm.get('canceled')
                        nm_stats['clicks'] = nm.get('clicks')
                        nm_stats['orders'] = nm.get('orders')
                        nm_stats['views'] = nm.get('views')
                        nm_stats['orderSum'] = nm.get('sum_price')
                        full_stats.append(nm_stats)
        full_stats_df = pd.DataFrame(full_stats, index=None)
        full_stats_df = full_stats_df.groupby('nmId').agg(
            sum = ('sum', 'sum'),
            atbs = ('atbs', 'sum'),
            canceled = ('canceled', 'sum'),
            clicks = ('clicks', 'sum'),
            orders = ('orders', 'sum'),
            views = ('views', 'sum'),
            sum_price = ('orderSum', 'sum'),
        ).reset_index()
        full_stats_df['orders'] = full_stats_df['orders'] - full_stats_df['canceled']
        full_stats_df = full_stats_df.drop(columns=['canceled'])
        adverts_stats = full_stats_df.to_dict(orient='records')
        return adverts_stats
    else:
        logger.success(f'Ошибка получения рекламной статистики за {date}')
        return None
            
### ------------------------------------------------- ### 

# В days передается кол-во дней для обновления отчета (21 дня достаточно, чтобы собрать все обновленные данные)
days = 21
# При помощи инструментов pandas собираем список дат 
dates = pd.date_range(end=pd.Timestamp.now()-relativedelta(days=1), periods=days, freq='D').strftime('%d.%m.%Y').tolist()
UPDATE_DAILY_DETAIL_HISTORY_REPORT(dates=dates, TOKEN_NAME='КОСТРИК')
